import openpyxl
import json
import os
import re
import requests
import cloudinary
import cloudinary.uploader
from constants import PAGE_METADATA, TABS_METADATA

# 🔧 Configure Cloudinary
cloudinary.config(
    cloud_name="dfncm107l",   # <-- Replace with yours
    api_key="238343496614657",         # <-- Replace with yours
    api_secret="wiv5t7TnY7cIgOFCkEHYP-GDWIo"    # <-- Replace with yours
)


def convert_drive_link_to_direct_url(link):
    if not isinstance(link, str):
        return ''
    match = re.search(r"/d/([a-zA-Z0-9_-]+)", link)
    if not match:
        match = re.search(r"id=([a-zA-Z0-9_-]+)", link)
    if match:
        file_id = match.group(1)
        return f"https://drive.google.com/uc?export=view&id={file_id}"
    return link.strip()


def download_image(file_id, save_path):
    url = f"https://drive.google.com/uc?export=download&id={file_id}"
    try:
        response = requests.get(url)
        if response.status_code == 200:
            with open(save_path, 'wb') as f:
                f.write(response.content)
            print(f"✅ Downloaded image to {save_path}")
            return True
        else:
            print(f"❌ Failed to download image for file ID {file_id}")
            return False
    except Exception as e:
        print(f"❌ Exception downloading image: {e}")
        return False


def upload_to_cloudinary(local_path, public_id):
    try:
        result = cloudinary.uploader.upload(local_path, public_id=public_id, folder="partners")
        # 🧹 Delete the file after successful upload
        os.remove(local_path)
        return result['secure_url']
    except Exception as e:
        print(f"❌ Cloudinary upload failed: {e}")
        # 🧹 Delete the file after successful upload
        os.remove(local_path)
        return ""


def excel_to_json_partners():
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(script_dir, "suma.xlsx")
        json_path = os.path.join(script_dir, "public/assets", "landing-page.json")
        network_data_path = os.path.join(script_dir, "public/assets", "network-data.json")
        images_dir = os.path.join(script_dir, "temp_downloads")
        os.makedirs(images_dir, exist_ok=True)

        workbook = openpyxl.load_workbook(file_path, data_only=True)
        try:
            sheet = workbook[PAGE_METADATA["PARTNERS"]]
        except KeyError:
            print(f"❌ Error: Sheet '{PAGE_METADATA['PARTNERS']}' not found.")
            print("Available sheets:", workbook.sheetnames)
            return

        headers = [str(cell.value).strip() if cell.value is not None else '' for cell in sheet[1]]
        expected_columns = TABS_METADATA["PARTNERS"]

        if not all(col in headers for col in expected_columns):
            print("❌ Error: Missing required columns.")
            print("Expected:", expected_columns)
            print("Found:", headers)
            return

        data = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                raw_name = row[headers.index(expected_columns[0])]
                if not raw_name or not str(raw_name).strip():
                    continue

                raw_src = row[headers.index(expected_columns[1])] or ''
                logo_url = convert_drive_link_to_direct_url(raw_src)

                file_match = re.search(r"id=([a-zA-Z0-9_-]+)", logo_url)
                if not file_match:
                    file_match = re.search(r"/d/([a-zA-Z0-9_-]+)", raw_src)
                file_id = file_match.group(1) if file_match else ''

                name_clean = str(raw_name).strip().lower()
                name_clean = re.sub(r'[^a-z0-9_-]', '', name_clean.replace(" ", "_"))
                local_filename = f"{name_clean}.jpg"
                local_path = os.path.join(images_dir, local_filename)

                cloudinary_url = ""
                if file_id:
                    if download_image(file_id, local_path):
                        cloudinary_url = upload_to_cloudinary(local_path, name_clean)
                    final_src = cloudinary_url or logo_url
                else:
                    final_src = logo_url

                row_data = {
                    'id': name_clean,
                    'src': final_src,
                    'alt': name_clean,
                    'name': str(raw_name).strip(),
                    'partnerState': row[headers.index(expected_columns[2])] or '',
                    'category': row[headers.index(expected_columns[3])] or '',
                    'website': row[headers.index(expected_columns[4])] or ''
                }
                data.append(row_data)

            except Exception as e:
                print(f"⚠️ Error processing row {row_idx}: {e}")
                continue

        if os.path.exists(json_path):
            with open(json_path, 'r', encoding='utf-8') as f:
                try:
                    json_data = json.load(f)
                except json.JSONDecodeError:
                    json_data = []
        else:
            json_data = []

        if not isinstance(json_data, list):
            json_data = [json_data]

        found = False
        for obj in json_data:
            if isinstance(obj, dict) and obj.get('type', '').strip().lower() == 'partner-logos':
                if 'partners' not in obj or not isinstance(obj['partners'], list):
                    obj['partners'] = []
                obj['partners'] = data
                found = True
                break

        if not found:
            json_data.append({
                "type": "partner-logos",
                "width": "100%",
                "position": "left",
                "title": "Our Network",
                "showFilters": False,
                "partners": data,
                "styles": {
                    "section": "partner-logos-section",
                    "title": "section-title",
                    "category": "partner-category",
                    "logosContainer": "logos-container",
                    "logo": "partner-logo"
                }
            })

        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, indent=2, ensure_ascii=False)

        print("✅ landing-page.json updated.")

        if os.path.exists(network_data_path):
            with open(network_data_path, 'r', encoding='utf-8') as f:
                try:
                    network_data = json.load(f)
                except json.JSONDecodeError:
                    network_data = {}
        else:
            network_data = {}

        if 'partners' not in network_data or not isinstance(network_data['partners'], list):
            network_data['partners'] = []

        existing_ids = {p['id'] for p in network_data['partners'] if 'id' in p}
        new_partners = [p for p in data if p['id'] not in existing_ids]

        if new_partners:
            network_data['partners'].extend(new_partners)
            with open(network_data_path, 'w', encoding='utf-8') as f:
                json.dump(network_data, f, indent=2, ensure_ascii=False)
            print(f"✅ {len(new_partners)} new partners added to network-data.json.")
        else:
            print("ℹ️ No new partners to add to network-data.json.")

    except Exception as e:
        print(f"❌ Unexpected error: {e}")


if __name__ == "__main__":
    excel_to_json_partners()
