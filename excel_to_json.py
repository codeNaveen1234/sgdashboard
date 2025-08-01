import openpyxl
import json
import os

def excel_to_json():
    try:
        # Get the directory of the current script
        script_dir = os.path.dirname(os.path.abspath(__file__))
        print(f"Looking for doc.xlsx in directory: {script_dir}")
        # Define the path to the Excel file
        file_path = os.path.join(script_dir, "doc.xlsx")
        # Define the path to the JSON file
        json_path = os.path.join(script_dir, "public/assets", "landing-page.json")

        # Verify Excel file exists
        if not os.path.exists(file_path):
            print(f"Error: File 'doc.xlsx' not found in directory: {script_dir}")
            return

        # Verify JSON file exists
        if not os.path.exists(json_path):
            print(f"Error: JSON file 'data.json' not found at: {json_path}")
            return

        # Open the Excel file
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        try:
            sheet = workbook["Data on homepage"]
        except KeyError:
            print("Error: Sheet 'Data on homepage' not found in the Excel file.")
            print(f"Available sheets: {workbook.sheetnames}")
            return

        # Print sheet details
        print(f"Sheet 'Data on homepage' found. Rows: {sheet.max_row}, Columns: {sheet.max_column}")

        # Get headers
        headers = [cell.value for cell in sheet[1]]
        cleaned_headers = [str(cell).strip() if cell is not None else '' for cell in headers]
        expected_columns = ['Indicator', 'Definition', 'Data', 'Icon link']
        if not all(col in cleaned_headers for col in expected_columns):
            print(f"Error: Excel file must contain columns: {expected_columns}")
            print(f"Found: {cleaned_headers}")
            return

        # Extract data rows
        data = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                row_data = {
                    'label': row[cleaned_headers.index('Indicator')] or '',
                    'value': row[cleaned_headers.index('Data')] or '',
                    'icon': row[cleaned_headers.index('Icon link')] or ''
                }
                if isinstance(row_data['value'], float) and row_data['value'].is_integer():
                    row_data['value'] = int(row_data['value'])
                data.append(row_data)
            except Exception as e:
                print(f"Error processing row {row_idx}: {str(e)}")
                continue

        print(f"Extracted {len(data)} rows of data: {data}")

        # Read the existing JSON file
        with open(json_path, 'r', encoding='utf-8') as json_file:
            raw_content = json_file.read()
            print(f"Raw content of {json_path}:")
            print(raw_content)
            json_file.seek(0)
            try:
                json_data = json.load(json_file)
                print("Parsed JSON data:", json_data)
            except json.JSONDecodeError:
                print("Warning: Invalid JSON. Attempting to parse as string.")
                try:
                    json_data = json.loads(raw_content)
                    print("Parsed JSON string:", json_data)
                except json.JSONDecodeError:
                    print("Error: Invalid JSON string. Using empty list.")
                    json_data = []

        # Ensure json_data is a list
        if not isinstance(json_data, list):
            print(f"Warning: JSON data is not a list (type: {type(json_data)}). Wrapping in a list.")
            json_data = [json_data] if json_data else []

        # Debug: Print all object types
        print("Objects in JSON array:")
        for obj in json_data:
            if isinstance(obj, dict):
                type_value = obj.get('type', 'N/A').strip()
                print(f" - Type: '{type_value}' (original: '{obj.get('type', 'N/A')}')")
            else:
                print(f" - Non-dictionary object: {obj}")

        # Find and update the "data-indicators" object
        found = False
        for obj in json_data:
            if isinstance(obj, dict) and obj.get('type', '').strip().lower() == 'data-indicators':
                print(f"Found 'data-indicators' object: {obj}")
                obj['indicators'] = data
                found = True
                break

        # If not found, append (this is a fallback, but your JSON should already have it)
        if not found:
            print("No 'data-indicators' object found. Appending new object (this should not happen if it exists).")
            json_data.append({'type': 'data-indicators', 'indicators': data})

        print("Updated JSON data:", json_data)

        # Write back to the file
        with open(json_path, 'w', encoding='utf-8') as json_file:
            json.dump(json_data, json_file, indent=2, ensure_ascii=False)

        print(f"JSON file updated at: {json_path} with {len(data)} rows.")

    except Exception as e:
        print(f"Error: {str(e)}")

if __name__ == "__main__":
    excel_to_json()
